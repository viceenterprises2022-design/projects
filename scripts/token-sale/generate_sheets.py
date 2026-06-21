import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_model():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styling helper variables
    font_family = "Calibri"
    
    # Fonts
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=13, bold=True, color="1F4E78")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=11, bold=True)
    font_italic = Font(name=font_family, size=9, italic=True, color="595959")
    font_regular = Font(name=font_family, size=11)
    
    # Fills
    fill_navy_dark = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_navy_light = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_accent = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_green_light = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_gray_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_title = Alignment(horizontal="left", vertical="center", indent=1)

    # Borders
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    border_double_bottom_side = Side(style='double', color='000000')
    border_thin_top_side = Side(style='thin', color='000000')
    border_total = Border(top=border_thin_top_side, bottom=border_double_bottom_side)

    # Helper function to auto-adjust column widths
    def autofit_columns(ws, min_width=12, padding=3):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='): # skip formulas for length
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)

    # -------------------------------------------------------------
    # SHEET 1: DASHBOARD & EXECUTIVE SUMMARY
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Dashboard & Summary")
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A1:G2")
    ws1["A1"] = "OMEGA V1 PROXY MODEL - TOKEN SALE BUSINESS CASE"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_navy_dark
    ws1["A1"].alignment = align_title
    ws1.row_dimensions[1].height = 25
    ws1.row_dimensions[2].height = 20

    # Meta Info
    ws1["A4"] = "Landing Cost (per 1M Tokens):"
    ws1["A4"].font = font_bold
    ws1["B4"] = 0.11
    ws1["B4"].number_format = "$#,##0.00"
    ws1["B4"].font = font_regular
    
    ws1["A5"] = "Standard Discount Offered:"
    ws1["A5"].font = font_bold
    ws1["B5"] = 0.20
    ws1["B5"].number_format = "0.0%"
    ws1["B5"].font = font_regular

    # Section: Key Metrics Dashboard (KPI Cards)
    ws1["A7"] = "Key Strategic Targets & Metrics"
    ws1["A7"].font = font_section
    
    # KPI 1: Landing Cost per 1B pack
    ws1.merge_cells("A8:B8")
    ws1["A8"] = "1B Pack Landing Cost"
    ws1["A8"].font = font_bold
    ws1["A8"].fill = fill_gray_light
    ws1["A8"].alignment = align_center
    
    ws1.merge_cells("A9:B9")
    ws1["A9"] = "=B4*1000"
    ws1["A9"].font = font_title
    ws1["A9"].font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    ws1["A9"].fill = fill_accent
    ws1["A9"].alignment = align_center
    ws1["A9"].number_format = "$#,##0.00"

    # KPI 2: Selling Price (Target Blended)
    ws1.merge_cells("C8:D8")
    ws1["C8"] = "1B Pack Selling Price (20% Disc.)"
    ws1["C8"].font = font_bold
    ws1["C8"].fill = fill_gray_light
    ws1["C8"].alignment = align_center
    
    ws1.merge_cells("C9:D9")
    ws1["C9"] = "='Pricing & Competitor Analysis'!E11*1000" # References dynamic cell in Sheet 2
    ws1["C9"].font = Font(name=font_family, size=16, bold=True, color="2E7D32")
    ws1["C9"].fill = fill_green_light
    ws1["C9"].alignment = align_center
    ws1["C9"].number_format = "$#,##0.00"

    # KPI 3: Gross Margin %
    ws1.merge_cells("E8:F8")
    ws1["E8"] = "Projected Gross Margin %"
    ws1["E8"].font = font_bold
    ws1["E8"].fill = fill_gray_light
    ws1["E8"].alignment = align_center
    
    ws1.merge_cells("E9:F9")
    ws1["E9"] = "=(C9-A9)/C9"
    ws1["E9"].font = Font(name=font_family, size=16, bold=True, color="2E7D32")
    ws1["E9"].fill = fill_green_light
    ws1["E9"].alignment = align_center
    ws1["E9"].number_format = "0.0%"

    # Strategy Brief Section
    ws1["A12"] = "Executive Business Case & Value Proposition"
    ws1["A12"].font = font_section
    
    strategy_points = [
        ("Core Product", "Omega V1 Proxy Framework - unified API access key to OpenAI (GPT-4o), Anthropic (Claude 3.5), & Gemini 1.5."),
        ("Value Proposition", "Startups/SMEs get a 20% discount compared to direct provider billing with zero expiration, higher limits, and no credit card lockups."),
        ("Cost Advantage", "Our landing cost of $0.11/M tokens is a fraction of market retail, generating >90% gross margins on premium blended traffic."),
        ("Target Market", "B2B customers, developers, SaaS startups, SMEs, and MSMEs building AI tools and automations."),
        ("Delivery Method", "Shared as a secure proxy key with private analytics, usage tracking, and automated balance warnings.")
    ]
    
    row_idx = 14
    for title, desc in strategy_points:
        ws1.cell(row=row_idx, column=1, value=title).font = font_bold
        ws1.cell(row=row_idx, column=1).alignment = align_left
        ws1.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
        ws1.cell(row=row_idx, column=2, value=desc).font = font_regular
        ws1.cell(row=row_idx, column=2).alignment = align_left
        row_idx += 1

    autofit_columns(ws1, min_width=15)

    # -------------------------------------------------------------
    # SHEET 2: PRICING & COMPETITOR ANALYSIS
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Pricing & Competitor Analysis")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:G2")
    ws2["A1"] = "PROVIDER COST COMPARISON & RESELLER PRICING MODEL"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_navy_dark
    ws2["A1"].alignment = align_title
    
    # Headers
    headers_s2 = ["AI Model / Provider", "Direct Input ($/1M)", "Direct Output ($/1M)", "Blended Retail Cost ($/1M)*", "Omega Standard ($/1M)", "Omega 20% Disc. ($/1M)", "Gross Margin / 1M"]
    for col_idx, h in enumerate(headers_s2, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy_light
        cell.alignment = align_center
    ws2.row_dimensions[4].height = 28

    # Input Data
    providers_data = [
        # Provider Name, Input, Output, Blended Weight (80% Input / 20% Output)
        ("OpenAI GPT-4o", 5.00, 15.00, "=B5*0.8+C5*0.2", 3.00, "=E5*(1-'Dashboard & Summary'!B5)", "=F5-'Dashboard & Summary'!B4"),
        ("Anthropic Claude 3.5 Sonnet", 3.00, 15.00, "=B6*0.8+C6*0.2", 2.50, "=E6*(1-'Dashboard & Summary'!B5)", "=F6-'Dashboard & Summary'!B4"),
        ("Gemini 1.5 Pro (Standard)", 1.25, 3.75, "=B7*0.8+C7*0.2", 1.00, "=E7*(1-'Dashboard & Summary'!B5)", "=F7-'Dashboard & Summary'!B4"),
        ("OpenAI GPT-4o mini", 0.150, 0.600, "=B8*0.8+C8*0.2", 0.20, "=E8*(1-'Dashboard & Summary'!B5)", "=F8-'Dashboard & Summary'!B4"),
        ("Gemini 1.5 Flash", 0.075, 0.300, "=B9*0.8+C9*0.2", 0.12, "=E9*(1-'Dashboard & Summary'!B5)", "=F9-'Dashboard & Summary'!B4"),
    ]

    for row_offset, row_data in enumerate(providers_data, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_offset, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border_all
            
            # Formatting rules
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx in [2, 3, 4, 5, 6, 7]:
                cell.alignment = align_right
                cell.number_format = "$#,##0.000" if "mini" in row_data[0] or "Flash" in row_data[0] else "$#,##0.00"

    # Dynamic Blended Weighted Average
    ws2["A11"] = "Target Hybrid Portfolio (Blended Average)"
    ws2["A11"].font = font_bold
    ws2["A11"].border = border_total
    
    ws2["B11"] = "=AVERAGE(B5:B7)"
    ws2["B11"].font = font_bold
    ws2["B11"].number_format = "$#,##0.00"
    ws2["B11"].border = border_total
    
    ws2["C11"] = "=AVERAGE(C5:C7)"
    ws2["C11"].font = font_bold
    ws2["C11"].number_format = "$#,##0.00"
    ws2["C11"].border = border_total

    ws2["D11"] = "=B11*0.8+C11*0.2"
    ws2["D11"].font = font_bold
    ws2["D11"].number_format = "$#,##0.00"
    ws2["D11"].border = border_total
    
    ws2["E11"] = "=AVERAGE(E5:E7)"
    ws2["E11"].font = font_bold
    ws2["E11"].number_format = "$#,##0.00"
    ws2["E11"].border = border_total

    ws2["F11"] = "=E11*(1-'Dashboard & Summary'!B5)"
    ws2["F11"].font = font_bold
    ws2["F11"].number_format = "$#,##0.00"
    ws2["F11"].fill = fill_green_light
    ws2["F11"].border = border_total

    ws2["G11"] = "=F11-'Dashboard & Summary'!B4"
    ws2["G11"].font = font_bold
    ws2["G11"].number_format = "$#,##0.00"
    ws2["G11"].fill = fill_green_light
    ws2["G11"].border = border_total

    # Note
    ws2["A13"] = "* Note: Blended Direct Cost assumes a typical 80% input tokens and 20% output tokens usage ratio."
    ws2["A13"].font = font_italic

    autofit_columns(ws2, min_width=15)

    # -------------------------------------------------------------
    # SHEET 3: FINANCIAL PROJECTIONS (12-MONTH MODEL)
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Financial Projections")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:N2")
    ws3["A1"] = "12-MONTH RESELLER REVENUE & MARGIN FORECAST"
    ws3["A1"].font = font_title
    ws3["A1"].fill = fill_navy_dark
    ws3["A1"].alignment = align_title
    
    # Months columns
    months_headers = ["Metric / Month", "Month 1", "Month 2", "Month 3", "Month 4", "Month 5", "Month 6", "Month 7", "Month 8", "Month 9", "Month 10", "Month 11", "Month 12", "Total / Avg"]
    for col_idx, m in enumerate(months_headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=m)
        cell.font = font_header
        cell.fill = fill_navy_light
        cell.alignment = align_center
    ws3.row_dimensions[4].height = 25

    # Rows definition
    # (Row Label, Font style, Cell Format, Formula/Values)
    proj_rows = [
        ("Active SME/Startup Clients", font_bold, "#,##0", [5, 8, 12, 18, 25, 32, 40, 48, 55, 62, 70, 80]),
        ("Average 1B Packs Sold / Client / Month", font_regular, "0.0", [1.0, 1.0, 1.1, 1.1, 1.2, 1.2, 1.3, 1.3, 1.4, 1.4, 1.5, 1.5]),
        ("Total 1B Packs Sold", font_bold, "#,##0", ["=B5*B6", "=C5*C6", "=D5*D6", "=E5*E6", "=F5*F6", "=G5*G6", "=H5*H6", "=I5*I6", "=J5*J6", "=K5*K6", "=L5*L6", "=M5*M6"]),
        ("Reseller Selling Price per 1B Pack", font_regular, "$#,##0", ["='Dashboard & Summary'!C9"]*12),
        
        # Revenue Section
        ("Gross Reseller Revenue", font_bold, "$#,##0", ["=B7*B8", "=C7*C8", "=D7*D8", "=E7*E8", "=F7*F8", "=G7*G8", "=H7*H8", "=I7*I8", "=J7*J8", "=K7*K8", "=L7*L8", "=M7*M8"]),
        
        # Cost Section
        ("Token Landing Cost (COGS)", font_regular, "$#,##0", ["=B7*'Dashboard & Summary'!A9"]*12),
        
        # Gross Margin Section
        ("Gross Margin Profit", font_bold, "$#,##0", ["=B9-B10", "=C9-C10", "=D9-D10", "=E9-E10", "=F9-F10", "=G9-G10", "=H9-H10", "=I9-I10", "=J9-J10", "=K9-K10", "=L9-L10", "=M9-M10"]),
        ("Gross Margin %", font_bold, "0.0%", ["=B11/B9", "=C11/C9", "=D11/D9", "=E11/E9", "=F11/F9", "=G11/G9", "=H11/H9", "=I11/I9", "=J11/J9", "=K11/K9", "=L11/L9", "=M11/M9"]),
        
        # Operational Costs
        ("Sales Marketing & CAC ($150/new client)", font_regular, "$#,##0", ["=150*B5", "=150*(C5-B5)", "=150*(D5-C5)", "=150*(E5-D5)", "=150*(F5-E5)", "=150*(G5-F5)", "=150*(H5-G5)", "=150*(I5-H5)", "=150*(J5-I5)", "=150*(K5-J5)", "=150*(L5-K5)", "=150*(M5-L5)"]),
        ("Proxy Platform hosting & APIs", font_regular, "$#,##0", [150, 150, 200, 200, 250, 250, 300, 300, 350, 350, 400, 400]),
        ("Total Operating Expenses", font_bold, "$#,##0", ["=SUM(B13:B14)", "=SUM(C13:C14)", "=SUM(D13:D14)", "=SUM(E13:E14)", "=SUM(F13:F14)", "=SUM(G13:G14)", "=SUM(H13:H14)", "=SUM(I13:I14)", "=SUM(J13:J14)", "=SUM(K13:K14)", "=SUM(L13:L14)", "=SUM(M13:M14)"]),
        
        # Net Profit Section
        ("Net Operating Profit", font_bold, "$#,##0", ["=B11-B15", "=C11-C15", "=D11-D15", "=E11-E15", "=F11-F15", "=G11-G15", "=H11-H15", "=I11-I15", "=J11-J15", "=K11-K15", "=L11-L15", "=M11-M15"])
    ]

    for idx, (label, font_style, fmt, values) in enumerate(proj_rows, 5):
        ws3.cell(row=idx, column=1, value=label).font = font_style
        ws3.cell(row=idx, column=1).alignment = align_left
        
        # Write Monthly Values
        for month_col, val in enumerate(values, 2):
            cell = ws3.cell(row=idx, column=month_col, value=val)
            cell.font = font_regular if font_style != font_bold else font_bold
            cell.number_format = fmt
            cell.alignment = align_right
            cell.border = border_all
            
            # Accent color for profits and margins
            if idx in [11, 12, 16] and font_style == font_bold:
                cell.fill = fill_green_light
            elif idx in [9, 15]:
                cell.fill = fill_accent

        # Write Totals/Averages Column
        tot_cell = ws3.cell(row=idx, column=14)
        tot_cell.font = font_bold
        tot_cell.alignment = align_right
        tot_cell.border = border_total
        col_letters = [get_column_letter(x) for x in range(2, 14)]
        
        if idx in [5, 6, 12]: # Average for counts, ratios, margin%
            tot_cell.value = f"=AVERAGE(B{idx}:M{idx})"
            tot_cell.number_format = fmt
        elif idx in [7, 9, 10, 11, 13, 14, 15, 16]: # Sum for financial items
            tot_cell.value = f"=SUM(B{idx}:M{idx})"
            tot_cell.number_format = fmt
            tot_cell.fill = fill_green_light if idx in [11, 16] else fill_accent
        elif idx == 8: # Blended Retail Cost - reference sheet 2
            tot_cell.value = "='Dashboard & Summary'!C9"
            tot_cell.number_format = fmt

    autofit_columns(ws3, min_width=12)

    # -------------------------------------------------------------
    # SHEET 4: TARGET CUSTOMER SEGMENTATION
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Customer Segmentation")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.merge_cells("A1:G2")
    ws4["A1"] = "TARGET B2B SEGMENTS & PRICING PLAN"
    ws4["A1"].font = font_title
    ws4["A1"].fill = fill_navy_dark
    ws4["A1"].alignment = align_title
    
    # Headers
    headers_s4 = ["Customer Segment", "Ideal Use Case", "Target Monthly Vol.", "Standard Price ($/1M)", "Omega Disc. ($/1M)", "Omega 1B Pack Price", "Expected Gross Margin"]
    for col_idx, h in enumerate(headers_s4, 1):
        cell = ws4.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy_light
        cell.alignment = align_center
    ws4.row_dimensions[4].height = 25

    segments_data = [
        ("Micro/SMEs (MSMEs)", "Internal Slackbots, automated email responses, basic CRM summaries.", "10M - 50M tokens", 2.20, 1.76, "=E5*1000", "=F5-('Dashboard & Summary'!B4*1000)"),
        ("SMEs (Small Enterprise)", "Customer support agents, automated code checking, content gen.", "100M - 500M tokens", 1.80, 1.44, "=E6*1000", "=F6-('Dashboard & Summary'!B4*1000)"),
        ("Fast-Growing Startups", "AI-native SaaS tools, LLM fine-tuning pipelines, agent workflows.", "500M - 2B tokens", 1.50, 1.20, "=E7*1000", "=F7-('Dashboard & Summary'!B4*1000)"),
        ("Enterprise / Companies", "Corporate knowledge management, secure local data processing.", "2B - 10B+ tokens", 1.20, 0.96, "=E8*1000", "=F8-('Dashboard & Summary'!B4*1000)"),
    ]

    for row_offset, row_data in enumerate(segments_data, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row_offset, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border_all
            
            if col_idx in [1, 2]:
                cell.alignment = align_left
            elif col_idx == 3:
                cell.alignment = align_center
            elif col_idx in [4, 5]:
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            elif col_idx in [6, 7]:
                cell.alignment = align_right
                cell.number_format = "$#,##0"
                cell.fill = fill_green_light if col_idx == 7 else fill_accent

    # Strategy Section below table
    ws4["A11"] = "Strategic GTM & Selling Plan"
    ws4["A11"].font = font_section
    
    gtm_points = [
        ("Startups Value Proposition", "Zero upfront API deposit + unified developer dashboard. Give them 20% discount on blended costs compared to OpenAI standard rates."),
        ("SME/MSME Value Proposition", "Non-technical setup. We provide pre-built endpoints and templates (e.g., Slackbots/Email replies) so they don't have to hire expensive AI engineers."),
        ("Sales Channel 1", "Direct outreach to early-stage SaaS companies via YC directory, Product Hunt, and TechCrunch launches."),
        ("Sales Channel 2", "Partnership with startup incubators & accelerators to offer 'Token Credits' as a member perk."),
        ("Customer Retention", "Implement sliding-scale volume discounts: as usage grows, token pricing drops, making client migration to direct keys unappealing.")
    ]
    
    row_idx = 13
    for title, desc in gtm_points:
        ws4.cell(row=row_idx, column=1, value=title).font = font_bold
        ws4.cell(row=row_idx, column=1).alignment = align_left
        ws4.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
        ws4.cell(row=row_idx, column=2, value=desc).font = font_regular
        ws4.cell(row=row_idx, column=2).alignment = align_left
        row_idx += 1

    autofit_columns(ws4, min_width=15)

    # -------------------------------------------------------------
    # SHEET 5: INTERACTIVE SALES & MARGIN CALCULATOR
    # -------------------------------------------------------------
    ws5 = wb.create_sheet(title="Sales Calculator")
    ws5.views.sheetView[0].showGridLines = True
    
    ws5.merge_cells("A1:D2")
    ws5["A1"] = "INTERACTIVE SALES & MARGIN CALCULATOR"
    ws5["A1"].font = font_title
    ws5["A1"].fill = fill_navy_dark
    ws5["A1"].alignment = align_title
    
    ws5["A4"] = "Enter business parameters below to calculate dynamic profitability."
    ws5["A4"].font = font_italic

    # Inputs Block
    ws5["A6"] = "Inputs (Editable)"
    ws5["A6"].font = font_bold
    ws5["A6"].fill = fill_accent
    
    ws5["A7"] = "Number of Active Clients:"
    ws5["A7"].font = font_regular
    ws5["B7"] = 10
    ws5["B7"].font = font_bold
    ws5["B7"].number_format = "#,##0"
    
    ws5["A8"] = "Average 1B Packs Sold/Client/Month:"
    ws5["A8"].font = font_regular
    ws5["B8"] = 1.5
    ws5["B8"].font = font_bold
    ws5["B8"].number_format = "0.0"

    ws5["A9"] = "Base Blended Retail Price ($/1M Tokens):"
    ws5["A9"].font = font_regular
    ws5["B9"] = 2.50
    ws5["B9"].font = font_bold
    ws5["B9"].number_format = "$#,##0.00"

    ws5["A10"] = "Discount Percentage Offered:"
    ws5["A10"].font = font_regular
    ws5["B10"] = 0.20
    ws5["B10"].font = font_bold
    ws5["B10"].number_format = "0.0%"

    # Outputs Block
    ws5["C6"] = "Calculated Outputs"
    ws5["C6"].font = font_bold
    ws5["C6"].fill = fill_accent
    
    ws5["C7"] = "Total 1B Packs Sold / Month:"
    ws5["C7"].font = font_regular
    ws5["D7"] = "=B7*B8"
    ws5["D7"].font = font_bold
    ws5["D7"].number_format = "#,##0.0"
    
    ws5["C8"] = "Selling Price per 1B Pack (with Discount):"
    ws5["C8"].font = font_regular
    ws5["D8"] = "=B9*(1-B10)*1000"
    ws5["D8"].font = font_bold
    ws5["D8"].number_format = "$#,##0.00"

    ws5["C9"] = "Total Monthly Reseller Revenue:"
    ws5["C9"].font = font_regular
    ws5["D9"] = "=D7*D8"
    ws5["D9"].font = font_bold
    ws5["D9"].number_format = "$#,##0.00"
    
    ws5["C10"] = "Total Monthly Landing Cost (COGS):"
    ws5["C10"].font = font_regular
    ws5["D10"] = "=D7*('Dashboard & Summary'!B4*1000)"
    ws5["D10"].font = font_bold
    ws5["D10"].number_format = "$#,##0.00"

    # Margin metrics
    ws5["C12"] = "Gross Monthly Profit:"
    ws5["C12"].font = font_bold
    ws5["C12"].fill = fill_green_light
    ws5["D12"] = "=D9-D10"
    ws5["D12"].font = font_bold
    ws5["D12"].number_format = "$#,##0.00"
    ws5["D12"].fill = fill_green_light
    ws5["D12"].border = border_total

    ws5["C13"] = "Gross Reseller Margin %:"
    ws5["C13"].font = font_bold
    ws5["C13"].fill = fill_green_light
    ws5["D13"] = "=D12/D9"
    ws5["D13"].font = font_bold
    ws5["D13"].number_format = "0.0%"
    ws5["D13"].fill = fill_green_light
    ws5["D13"].border = border_total

    # Add borders to Inputs/Outputs
    for r in range(6, 11):
        ws5.cell(row=r, column=1).border = border_all
        ws5.cell(row=r, column=2).border = border_all
        ws5.cell(row=r, column=3).border = border_all
        ws5.cell(row=r, column=4).border = border_all
        
    ws5.cell(row=12, column=3).border = border_all
    ws5.cell(row=12, column=4).border = border_all
    ws5.cell(row=13, column=3).border = border_all
    ws5.cell(row=13, column=4).border = border_all

    autofit_columns(ws5, min_width=25)

    # Save to file
    filepath = "/home/vreddy1/Desktop/Projects/scripts/token-sale/token_sale_model.xlsx"
    wb.save(filepath)
    print(f"Excel model successfully generated at: {filepath}")

if __name__ == "__main__":
    build_excel_model()
